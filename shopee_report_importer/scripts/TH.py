# driver_run_with_cookies_excel.py
"""
Cách dùng:
- Tạo file Excel, ví dụ: cookies.xlsx
- Trong Sheet1, cột A (từ A1 trở xuống), mỗi dòng là 1 cookie header.
- Chỉnh SCRIPTS = [...] thành 7 file python của bạn.
- Chạy: python driver_run_with_cookies_excel.py
"""
import os
import sys
import json
import time
import subprocess
from typing import List
from openpyxl import load_workbook

DOMAIN = ".shopee.vn"
COOKIES_JSON = "cookies_shopee.json"   # file mà 7 script đang đọc (ghi đè mỗi bộ cookie)
EXCEL_PATH = "headers.xlsx"            # Excel chứa cookie headers
SCRIPTS = [
    "laban_local.py", "ads_cpc.py", "ads_live_local.py",
    "booking_local.py", "live_product.py", "video_product.py", "order_local.py",
]
# Chọn "sequential" hoặc "parallel"
RUN_MODE = "sequential"


def header_to_cookie_list(header: str, domain: str) -> List[dict]:
    jar = {}
    for part in header.split(";"):
        part = part.strip()
        if not part or "=" not in part:
            continue
        k, v = part.split("=", 1)
        jar[k] = v
    return [{
        "name": k,
        "value": v,
        "domain": domain,
        "path": "/",
        "secure": True
    } for k, v in jar.items()]


def save_cookies_json(cookies: List[dict], out_path: str):
    out_dir = os.path.dirname(out_path)
    if out_dir:
        os.makedirs(out_dir, exist_ok=True)
    with open(out_path, "w", encoding="utf-8") as f:
        json.dump(cookies, f, ensure_ascii=False, indent=2)


def run_scripts_sequential(scripts: List[str], cookie_path: str):
    for script in scripts:
        cmd = [sys.executable, script, "--cookies", cookie_path]
        print(f"[SEQUENTIAL] Running: {' '.join(cmd)}")
        try:
            subprocess.run(cmd, check=True)
            print(f"[SEQUENTIAL] Finished: {script}")
        except subprocess.CalledProcessError as e:
            print(f"[SEQUENTIAL][ERROR] {script} returned {e.returncode}; continuing...")


def run_scripts_parallel(scripts: List[str], cookie_path: str):
    procs = []
    for script in scripts:
        cmd = [sys.executable, script, "--cookies", cookie_path]
        print(f"[PARALLEL] Starting: {' '.join(cmd)}")
        p = subprocess.Popen(cmd)
        procs.append((script, p))

    for script, p in procs:
        ret = p.wait()
        if ret == 0:
            print(f"[PARALLEL] Finished OK: {script}")
        else:
            print(f"[PARALLEL][ERROR] {script} exited with code {ret}")


def read_headers_from_excel(path: str) -> List[str]:
    wb = load_workbook(path)
    ws = wb["Sheet1"]
    headers = []
    for row in ws["A"]:  # cột A
        if row.value and isinstance(row.value, str) and row.value.strip():
            headers.append(row.value.strip())
    return headers


def main():
    if not os.path.exists(EXCEL_PATH):
        print(f"Không tìm thấy {EXCEL_PATH}. Hãy tạo file Excel và thêm cookie headers vào Sheet1 cột A.")
        return

    headers = read_headers_from_excel(EXCEL_PATH)
    if not headers:
        print("Không có cookie header nào trong file Excel.")
        return

    for idx, header in enumerate(headers, start=1):
        print("\n" + "=" * 40)
        print(f"BỘ COOKIE #{idx} / {len(headers)}")
        cookies = header_to_cookie_list(header, DOMAIN)
        save_cookies_json(cookies, COOKIES_JSON)
        print(f"Đã ghi {COOKIES_JSON} ({len(cookies)} cookies).")

        start = time.time()
        if RUN_MODE == "sequential":
            run_scripts_sequential(SCRIPTS, COOKIES_JSON)
        elif RUN_MODE == "parallel":
            run_scripts_parallel(SCRIPTS, COOKIES_JSON)
        else:
            print("RUN_MODE không hợp lệ:", RUN_MODE)
            return
        elapsed = time.time() - start
        print(f"Tất cả scripts cho BỘ #{idx} đã xong (thời gian {elapsed:.1f}s).")

        time.sleep(1)

    print("\nHoàn tất tất cả bộ cookie.")


if __name__ == "__main__":
    main()
